from django.shortcuts import render
import os
from django.shortcuts import render, redirect
from .forms import UploadFileForm
from openpyxl import load_workbook
import requests
from urllib.parse import urlencode
from django.http import HttpResponse, Http404
from .models import AddressFile
from django.contrib import messages
from django.conf import settings



def lat_long(address):
    params = {"location":address, "key": settings.API_KEY}
    url = f"{'http://www.mapquestapi.com/geocoding/v1/address'}?{urlencode(params)}"
    response = requests.get(url)
    try:
        lat = response.json()['results'][0]['locations'][0]["latLng"]["lat"]
        lon = response.json()['results'][0]['locations'][0]["latLng"]["lng"]
    except:
        pass   
    return lat,lon


def upload(request):
    if request.method == "POST":
        form = UploadFileForm(request.POST, request.FILES)
        if form.is_valid():
            file_name = request.FILES['excel_file']
            if not str(file_name).endswith('xlsx'):
                messages.info(request, 'Incorrect File Format. Please Upload .xlsx file')
                return redirect('/upload')
            else:
                wb = load_workbook(filename=file_name.file)
                for ws in wb.worksheets:
                    for index,row in enumerate(ws.rows,start=1):
                        if type(row[0].value) == str:
                            lat,lon = lat_long(row[0].value)
                            ws.cell(row=index,column=2).value = lat
                            ws.cell(row=index,column=3).value = lon
                        else:
                            continue
                wb.save(filename=file_name.file)
                form.save()
                return redirect('/upload')
    else:
        form = UploadFileForm()
    uploaded_file = AddressFile.objects.all()
    file_obj = uploaded_file.last()
    context = {
        'form': form,
        'file_obj': file_obj
    }
    return render(request, 'upload_form.html',context)
